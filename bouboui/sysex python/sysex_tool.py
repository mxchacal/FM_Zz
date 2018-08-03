from openpyxl import load_workbook
from time import sleep

print("------- DX7 Sysex Python Tool --------")

block_size = 1

#material to work with
filename = input("Sysex file : ")
print('Opening .syx ...')
sysex = open(filename, "rb")
header_file = open('cartridge.h', "w")
xls = load_workbook('cartridge.xlsx')
sheet = xls.active

sheet.title = "DX7 ROM CARTRIDGE"
sheet['B1'] = filename

block = 0
header = ''

#check / write header
print('Checking sysex header')
print('Header bytes : ', end = "", flush = True)
for i in range(0,6):
	block = sysex.read(block_size)
	sheet.cell(row = 2, column = i + 2).value = str(block.hex())
	header = header + str(block.hex()) + ' '
	print(str(block.hex()) + " ", end = "", flush = True)
	sleep(0.2)

if header == 'f0 43 00 09 20 00 ':
	print('\nHeader', end = "", flush = True)
	sleep(1)
	print(' OK')
	sleep(1)

print('Converting voices...')
sleep(1)

name = ["" for x in range(32)]
row_i = 4
index = 124

for j in range(0, 31):
	sysex.seek(index)
	for i in range(0, 10):
		name[j] = name[j] + str(chr(int.from_bytes(sysex.read(block_size), byteorder='big')))

	index = index + 128
	print(name[j] + " --> " + str(j + 1) + "/32 converted", end = "\r")
	sheet.cell(row = row_i, column = 2).value = name[j]
	row_i = row_i + 11
	sleep(0.1)

print(name[31] + "32/32 converted                      ")
print("Done!")

#write voice name

xls.save("cartridge.xlsx")